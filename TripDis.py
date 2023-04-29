import numpy as np
import math
import pandas as pd
from openpyxl import load_workbook

# The coordinates of the central locations chosen for each zone: list: [X,Y]

# z77 = [293861, 6197646] #Melun
# z75 = [261410, 6250947] #Paris
# z78 = [236003, 6239814] #Versailles
# z91 = [259498, 6220830] #Savigny sur Orge
# z92 = [250648, 6260134] #Asnières-sur-Seine
# z93 = [277530, 6258441] #Bondy
# z94 = [273614, 6237274] #Créteil
# z95 = [267052, 6274422] #Sarcelles

# coord = np.array([[293861, 6197646],
# [261410, 6250947],
# [236003, 6239814],
# [259498, 6220830],
# [250648, 6260134],
# [277530, 6258441],
# [273614, 6237274],
# [267052, 6274422]])

# Distance_array = np.zeros((8, 8))

# #To generate an origin destination matrix, all possible distances of zones with respect to each other have to be stored in an 8x8 matrix
# for i in range(8):
#     for j in range(8):

#         Dist = math.sqrt(math.pow(coord[i ,0]-coord[j,0],2)    + math.pow(coord[i,1]-coord[j,1],2))

#         Distance_array[i, j] =Dist

# print(Distance_array)


data = load_workbook(filename='distancematrix.xlsx',
                     data_only=True)
# data = pd.read_excel (r'distancematrix.xlsx')
ws = data['Sheet1']

# Read the cell values into a list of lists
# print("Crowfly distances")
data_rows1 = []
for row in ws['A1':'H8']:
    data_cols1 = []
    for cell in row:
        data_cols1.append(cell.value)
    data_rows1.append(data_cols1)
df1 = pd.DataFrame(data_rows1, index=list(["77", "75", "78", "91", "92", "93", "94", "95"]),
                   columns=list(["77", "75", "78", "91", "92", "93", "94", "95"]))

# print(df1)
# print()

# print("Travel time car [min]")
data_rows2 = []
for row in ws['B13':'I20']:
    data_cols2 = []
    for cell in row:
        data_cols2.append(cell.value)
    data_rows2.append(data_cols2)
df2 = pd.DataFrame(data_rows2, index=list(["77", "75", "78", "91", "92", "93", "94", "95"]),
                   columns=list(["77", "75", "78", "91", "92", "93", "94", "95"]))

# print(df2)
# print()

# print("Travel time PT [min]")
data_rows3 = []
for row in ws['B24':'I31']:
    data_cols3 = []
    for cell in row:
        data_cols3.append(cell.value)
    data_rows3.append(data_cols3)
df3 = pd.DataFrame(data_rows3, index=list(["77", "75", "78", "91", "92", "93", "94", "95"]),
                   columns=list(["77", "75", "78", "91", "92", "93", "94", "95"]))

# print(df3)
# print()

# print("Travel time UAM [min]")
data_rows4 = []
for row in ws['B35':'I42']:
    data_cols4 = []
    for cell in row:
        data_cols4.append(cell.value)
    data_rows4.append(data_cols4)
df4 = pd.DataFrame(data_rows4, index=list(["77", "75", "78", "91", "92", "93", "94", "95"]),
                   columns=list(["77", "75", "78", "91", "92", "93", "94", "95"]))

# print(df4)
# print()

# print("Cost car [euro]")
data_rows5 = []
for row in ws['B46':'I53']:
    data_cols5 = []
    for cell in row:
        data_cols5.append(cell.value)
    data_rows5.append(data_cols5)
df5 = pd.DataFrame(data_rows5, index=list(["77", "75", "78", "91", "92", "93", "94", "95"]),
                   columns=list(["77", "75", "78", "91", "92", "93", "94", "95"]))

# print(df5)
# print()
#
# print("Cost PT [euro]")
data_rows6 = []
for row in ws['B57':'I64']:
    data_cols6 = []
    for cell in row:
        data_cols6.append(cell.value)
    data_rows6.append(data_cols6)
df6 = pd.DataFrame(data_rows6, index=list(["77", "75", "78", "91", "92", "93", "94", "95"]),
                   columns=list(["77", "75", "78", "91", "92", "93", "94", "95"]))

# print(df6)
# print()

# print("Cost UAM [euro]")
data_rows7 = []
for row in ws['B68':'I75']:
    data_cols7 = []
    for cell in row:
        data_cols7.append(cell.value)
    data_rows7.append(data_cols7)
df7 = pd.DataFrame(data_rows7, index=list(["77", "75", "78", "91", "92", "93", "94", "95"]),
                   columns=list(["77", "75", "78", "91", "92", "93", "94", "95"]))

# print(df7)
# print()